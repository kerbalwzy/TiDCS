# -*- coding:utf-8 -*-
# @Author: wzy
# 创建任务的相关接口
from datetime import datetime
from typing import Iterator, Dict, Any

from flask import request
from flask_restful import reqparse
from openpyxl import Workbook, load_workbook

from backend.consts import RespCode
from backend.http_apis import httpApi, json_resp, excel_resp
from backend.models import Product, db, query_paginate

list_product_parser = reqparse.RequestParser(trim=True)
list_product_parser.add_argument("page", type=int, location="args", default=1)
list_product_parser.add_argument("limit", type=int, location="args", default=0)
list_product_parser.add_argument("code", type=str, location="args")
list_product_parser.add_argument("desc", type=str, location="args")


@httpApi.route("/products", methods=["GET"], endpoint="获取产品列表")
def list_products():
    params = list_product_parser.parse_args()
    filters = [Product.deleteTime.is_(None)]
    if params["code"]:
        filters.append(Product.code.like(f"%{params['code']}%"))
    if params["desc"]:
        filters.append(Product.desc.like(f"%{params['desc']}%"))
    query = db.session.query(Product).filter(*filters).order_by(Product.createTime.desc()).order_by(Product.code.asc())
    result, count = query_paginate(query=query, page=params["page"], limit=params["limit"])
    # 解析数据, 并检查数据
    data = list()
    for i in result:
        data.append(i.to_dict())
    return json_resp(errcode=0, msg="success", data=data, count=count)


@httpApi.route("/products/xlsx/template", methods=["GET"], endpoint="下载产品上传Excel模版文件")
def download_products_xlsx_template():
    wb = Workbook()
    ws = wb.active
    ws.append(["产品型号", "抢购上限"])
    return excel_resp(wb=wb, filename="TI产品抢购.xlsx")


def parse_excel_row(row: Iterator, row_index: int, index_header_map: Dict[int, str]) -> Dict[str, Any]:
    # 提取数据
    temp = dict()
    for index, col in enumerate(row):
        cn_key = index_header_map.get(index)
        if cn_key not in ["产品型号", "抢购上限"]:
            continue
        if cn_key == "产品型号":
            assert isinstance(col.value, str), ValueError(f"数据列[产品型号]， 第{row_index}行, 要求数据类型为文本字符串")
            temp[cn_key] = col.value.strip()
        if cn_key == "抢购上限":
            assert isinstance(col.value, int) or isinstance(col.value, float), ValueError(
                f"数据列[产品型号]， 第{row_index}行, 要求数据类型为数值")
            temp[cn_key] = int(col.value)
    return temp


@httpApi.route("/products", methods=["POST"], endpoint="上传TI产品抢购Excel文件")
def upload_ti_products_excel():
    file_fp = request.files.get("file")
    if not file_fp:
        return json_resp(errcode=RespCode.ParamsError, msg="请选择文件上传")
    # 读取文件
    wb = load_workbook(filename=file_fp)
    ws = wb.active
    _rows = ws.rows
    # 文件表头检查
    index_header_map = {index: c.value for index, c in enumerate(next(_rows))}  # 第一行
    if {"产品型号", "抢购上限"}.issubset(set(index_header_map.values())) is False:
        wb.close()
        return json_resp(errcode=RespCode.ParamsError, msg=f"上传文件中缺少必要的列名, 请按照模板文件填充数据")
    # 文件内容数据检查
    data = dict()
    row_index = 1
    # 实际是从第二行(index=1)开始遍历的, 因为之前已经next过了
    for r in _rows:
        row_index += 1
        try:
            item = parse_excel_row(row=r, row_index=row_index, index_header_map=index_header_map)
        except Exception as e:
            return json_resp(errcode=RespCode.ParamsError, msg=e.__str__())
        else:
            data[item["产品型号"]] = item["抢购上限"]
    # 查询并更新已经存在的产品
    products = Product.query.filter(Product.code.in_(list(data.keys()))).all()
    for item in products:
        if item.deleteTime is not None:
            item.deleteTime = None
            item.createTime = datetime.utcnow()
            item.updateTime = datetime.utcnow()
        item.wantLimit = data.pop(item.code)
        db.session.add(item)
    # 新增不存在的产品
    for k, v in data.items():
        item = Product(code=k, wantLimit=v)
        db.session.add(item)
    # 提交数据
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise e
    return json_resp(errcode=0, msg="success")


del_products_parser = reqparse.RequestParser()
del_products_parser.add_argument("codes", type=list, required=True, location="json")


@httpApi.route("/products", methods=["DELETE"], endpoint="批量删除产品记录")
def multi_delete_products():
    params = del_products_parser.parse_args()
    if len(params["codes"]) == 0:
        return json_resp(errcode=0, msg="success")
    products = Product.un_delete_query().filter(Product.code.in_(params["codes"])).all()
    for item in products:
        item.deleteTime = datetime.utcnow()
        db.session.add(item)
    # 提交数据
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise e
    return json_resp(errcode=0, msg="success")


update_products_parser = reqparse.RequestParser(trim=True)
update_products_parser.add_argument("code", type=str, required=True, location="json")
update_products_parser.add_argument("wantLimit", type=int, required=True, location="json")


@httpApi.route("/products", methods=["PUT"], endpoint="编辑单条产品记录")
def update_products():
    params = update_products_parser.parse_args()
    params["wantLimit"] = params["wantLimit"] if params["wantLimit"] > 0 else 0
    product = Product.un_delete_query().filter(Product.code == params["code"]).first()
    if not product:
        return json_resp(errcode=RespCode.ProductNotFound, msg="产品记录未找到")
    product.wantLimit = params["wantLimit"]
    db.session.add(product)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise e
    return json_resp(errcode=0, msg="success")
