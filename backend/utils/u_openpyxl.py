# -*- coding:utf-8 -*-
# @Author: wzy
#
import gc
from typing import List, Iterable

from openpyxl import Workbook


def excel_work_book(data: List[dict]) -> Workbook:
    """
    data:
    [
        {
            "sheet": <str> ,
            "title": <Iterable>,
            "content": <List[<Iterable>]>
        }
    ]
    """
    wb = Workbook(write_only=True)  # 开启write_only=True节省内存, 否则很容易内存溢出
    sheet_index = 0
    for item in data:
        # 单页数据小于100W条时正常保存
        if len(item["content"]) <= 1000000:
            ws = wb.create_sheet(title=item["sheet"], index=sheet_index)
            ws.append(item["title"])
            for row in item["content"]:
                ws.append(row)
            sheet_index += 1
        # 单页数据超过100W条时,需要分多页保存
        else:
            loop_count = 0
            while True:
                loop_count += 1
                temp_content = item["content"][(loop_count - 1) * 1000000:loop_count * 1000000]
                if not temp_content:
                    break
                ws = wb.create_sheet(title=f"{item['sheet']}-{loop_count}", index=sheet_index)
                ws.append(item["title"])
                for row in temp_content:
                    ws.append(row)
                sheet_index += 1
                del temp_content
        del item
    del data
    gc.collect()
    return wb


def make_sheet_content(title: Iterable, data: List[dict]) -> List[list]:
    res = list()
    for item in data:
        temp = list()
        for k in title:
            temp.append(item.get(k))
        res.append(temp)
    return res
